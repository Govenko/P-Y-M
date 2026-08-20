#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Standalone desktop HH.ru vacancy parser that drives Google Chrome.

Поиск вакансий на hh.ru идет через обычный Chrome (Playwright), потому что
анонимный доступ к API поиска hh.ru закрыт. Открытый справочник регионов
api.hh.ru/areas используется только для определения id региона.
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import os
import queue
import re
import subprocess
import sys
import threading
import time
import urllib.parse
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Callable, Iterable, Optional

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from playwright.async_api import TimeoutError as PlaywrightTimeoutError
from playwright.async_api import async_playwright


APP_TITLE = "HH.ru Parser"
ROOT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = ROOT_DIR / "results" / "hh"
DEFAULT_PROFILE_DIR = ROOT_DIR / "tools" / "hh_chrome_profile"

VACANCY_URL_RE = re.compile(r"https?://[^/]*hh\.ru/vacancy/(?P<id>\d+)")
EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)
PHONE_RE = re.compile(r"(?:\+7|8)\s*[\( -]?\d{3,5}[\) -]?\s*\d{1,3}[\s-]?\d{2,3}[\s-]?\d{2,3}")

BLOCK_MARKERS = (
    "подтвердите, что вы не робот",
    "доступ ограничен",
    "проверка браузера",
    "captcha",
)

# Запасной список, если каталог регионов не загрузится.
FALLBACK_AREAS = {
    "вся россия": "113",
    "москва": "1",
    "санкт-петербург": "2",
    "спб": "2",
    "екатеринбург": "3",
    "новосибирск": "4",
    "казань": "88",
    "нижний новгород": "66",
    "самара": "78",
    "уфа": "99",
    "пермь": "72",
    "челябинск": "104",
    "ростов-на-дону": "76",
    "краснодар": "53",
    "воронеж": "26",
    "тюмень": "95",
    "омск": "68",
    "красноярск": "54",
    "владивосток": "22",
    "ижевск": "96",
}

CITY_CHOICES = [
    "Вся Россия",
    "Москва",
    "Санкт-Петербург",
    "Казань",
    "Нижний Новгород",
    "Самара",
    "Уфа",
    "Екатеринбург",
    "Пермь",
    "Челябинск",
    "Ростов-на-Дону",
    "Краснодар",
    "Воронеж",
    "Новосибирск",
    "Тюмень",
    "Омск",
    "Красноярск",
    "Владивосток",
    "Ижевск",
]


@dataclass
class HHVacancy:
    title: str = ""
    salary: str = ""
    employer: str = ""
    employer_hh_url: str = ""
    area: str = ""
    contact_name: str = ""
    contact_email: str = ""
    contact_phones: str = ""
    description: str = ""
    published: str = ""
    url: str = ""
    source_query: str = ""
    source_city: str = ""
    collected_at: str = ""


def clean_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def unique(values: Iterable[str], limit: int = 20) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        clean = clean_text(value).strip(" ,;")
        key = clean.lower()
        if clean and key not in seen:
            seen.add(key)
            result.append(clean)
            if len(result) >= limit:
                break
    return result


def find_chrome() -> Path:
    candidates = [
        Path(os.environ.get("ProgramFiles", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
        Path(os.environ.get("ProgramFiles(x86)", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
        Path(os.environ.get("LOCALAPPDATA", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
    ]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError("Google Chrome не найден. Установите Chrome и запустите программу снова.")


def canonical_vacancy_url(url: str) -> str:
    match = VACANCY_URL_RE.search(url or "")
    if not match:
        return ""
    return f"https://hh.ru/vacancy/{match.group('id')}"


def vacancy_key(url: str) -> str:
    match = VACANCY_URL_RE.search(url or "")
    return match.group("id") if match else url


def flatten_areas(nodes: list[dict], into: dict[str, str]) -> dict[str, str]:
    for node in nodes or []:
        name = clean_text(node.get("name")).lower()
        if name and name not in into:
            into[name] = str(node.get("id") or "")
        flatten_areas(node.get("areas") or [], into)
    return into


def resolve_area(city: str, log: Callable[[str], None]) -> str:
    city_key = clean_text(city).lower().replace("ё", "е")
    if city_key in ("", "вся россия", "россия"):
        return "113"
    try:
        response = httpx.get(
            "https://api.hh.ru/areas",
            headers={"User-Agent": "HHLeadParser/1.0 (area lookup)"},
            timeout=20.0,
        )
        response.raise_for_status()
        areas = flatten_areas(response.json(), {})
        if city_key in areas:
            return areas[city_key]
    except Exception as exc:
        log(f"Каталог регионов не загрузился ({exc}), использую встроенный список.")
    if city_key in FALLBACK_AREAS:
        return FALLBACK_AREAS[city_key]
    log(f"Регион «{city}» не найден, ищу по всей России.")
    return "113"


def build_search_url(query: str, area: str, page_no: int) -> str:
    encoded = urllib.parse.quote(clean_text(query))
    url = f"https://hh.ru/search/vacancy?text={encoded}&area={area}&items_on_page=50"
    if page_no > 0:
        url += f"&page={page_no}"
    return url


async def safe_click_text(page, texts: Iterable[str], timeout: int = 1200) -> None:
    for text in texts:
        try:
            locator = page.get_by_text(text, exact=False).first
            if await locator.count():
                await locator.click(timeout=timeout)
                await page.wait_for_timeout(400)
                return
        except Exception:
            continue


async def wait_if_blocked(page, log: Callable[[str], None]) -> None:
    try:
        content = str(await page.evaluate("() => (document.body && document.body.innerText) || ''")).lower()
    except Exception:
        return
    if any(marker in content for marker in BLOCK_MARKERS):
        log("HH показал проверку (капча / доступ ограничен).")
        log("Решите проверку в открытом окне Chrome, программа подождет 90 секунд.")
        await page.wait_for_timeout(90_000)


async def extract_serp_items(page) -> list[dict]:
    return await page.evaluate(
        """
        () => {
          const pickText = (root, selectors) => {
            for (const selector of selectors) {
              const el = root.querySelector(selector);
              const text = el && (el.innerText || el.textContent || '').trim();
              if (text) return text;
            }
            return '';
          };
          let cards = Array.from(document.querySelectorAll(
            '[data-qa="vacancy-serp__vacancy"], [data-qa*="serp-item"], [class*="vacancy-card"]'
          ));
          if (!cards.length) {
            const links = Array.from(document.querySelectorAll('a[href*="/vacancy/"]'));
            const seen = new Set();
            cards = links.map(link => {
              let node = link;
              for (let i = 0; i < 6 && node.parentElement; i++) node = node.parentElement;
              return node;
            }).filter(node => {
              if (seen.has(node)) return false;
              seen.add(node);
              return true;
            });
          }
          return cards.map(card => {
            const link = card.querySelector('a[data-qa="serp-item__title"], a[href*="/vacancy/"]');
            return {
              title: pickText(card, ['[data-qa="serp-item__title"]', 'a[href*="/vacancy/"]', 'h2', 'h3']),
              href: link ? (link.href || '') : '',
              salary: pickText(card, [
                '[data-qa="vacancy-serp__vacancy-compensation"]',
                '[data-qa*="compensation"]',
                '[class*="compensation"]'
              ]),
              employer: pickText(card, [
                '[data-qa="vacancy-serp__vacancy-employer"]',
                'a[data-qa*="employer"]',
                '[class*="company-name"]'
              ]),
              employerHref: (card.querySelector('a[data-qa*="employer"], a[href*="/employer/"]') || {}).href || '',
              area: pickText(card, [
                '[data-qa="vacancy-serp__vacancy-address"]',
                '[data-qa*="address"]'
              ])
            };
          }).filter(item => item.href);
        }
        """
    )


async def enrich_from_vacancy_page(page, vacancy: HHVacancy, log: Callable[[str], None]) -> None:
    try:
        await page.goto(vacancy.url, wait_until="domcontentloaded", timeout=60_000)
        await page.wait_for_timeout(2200)
        await wait_if_blocked(page, log)
        await safe_click_text(page, ("Показать контакты", "Показать телефон"), timeout=900)
        await page.wait_for_timeout(800)
        data = await page.evaluate(
            """
            () => {
              const pick = selectors => {
                for (const selector of selectors) {
                  const el = document.querySelector(selector);
                  const text = el && (el.innerText || el.textContent || '').trim();
                  if (text) return text;
                }
                return '';
              };
              const mailto = Array.from(document.querySelectorAll('a[href^="mailto:"]'))
                .map(a => a.href.replace('mailto:', '').split('?')[0]);
              const tels = Array.from(document.querySelectorAll('a[href^="tel:"]'))
                .map(a => a.href.replace('tel:', ''));
              return {
                description: pick(['[data-qa="vacancy-description"]', '[class*="vacancy-description"]']),
                employer: pick(['a[data-qa="vacancy-company-name"]', '[data-qa*="company-name"]']),
                employerHref: (document.querySelector('a[data-qa="vacancy-company-name"], a[href*="/employer/"]') || {}).href || '',
                contactsBlock: pick(['[data-qa="vacancy-contacts"]', '[class*="vacancy-contacts"]']),
                contactName: pick(['[data-qa="vacancy-contacts__fio"]']),
                mailto,
                tels,
                published: pick(['[data-qa="vacancy-view-created"]', '[class*="vacancy-creation-time"]'])
              };
            }
            """
        )
        if clean_text(data.get("description")):
            vacancy.description = clean_text(data.get("description"))[:900]
        if clean_text(data.get("employer")):
            vacancy.employer = clean_text(data.get("employer"))
        if clean_text(data.get("employerHref")):
            vacancy.employer_hh_url = clean_text(data.get("employerHref")).split("?")[0]
        if clean_text(data.get("published")):
            vacancy.published = clean_text(data.get("published"))
        vacancy.contact_name = clean_text(data.get("contactName"))

        contacts_block = clean_text(data.get("contactsBlock"))
        emails = list(data.get("mailto") or []) + EMAIL_RE.findall(contacts_block)
        phones = list(data.get("tels") or []) + PHONE_RE.findall(contacts_block)
        vacancy.contact_email = "; ".join(unique(emails, limit=3))
        vacancy.contact_phones = "; ".join(unique(phones, limit=3))
    except PlaywrightTimeoutError:
        log("  ! таймаут вакансии, оставляю данные из выдачи")
    except Exception as exc:
        log(f"  ! ошибка вакансии: {exc}")


async def scrape_hh(
    *,
    query: str,
    city: str,
    max_results: int,
    max_pages: int,
    deep: bool,
    visible: bool,
    output_dir: Path,
    log: Callable[[str], None],
    item_callback: Optional[Callable[[HHVacancy], None]] = None,
) -> tuple[list[HHVacancy], dict[str, Path]]:
    chrome = find_chrome()
    DEFAULT_PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    area = resolve_area(city, log)

    log(f"Chrome: {chrome}")
    log(f"Регион HH: {city or 'Вся Россия'} (id={area})")
    log(f"Открываю hh.ru: {query}")
    vacancies: list[HHVacancy] = []
    seen: set[str] = set()

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            str(DEFAULT_PROFILE_DIR),
            executable_path=str(chrome),
            headless=not visible,
            viewport={"width": 1360, "height": 900},
            locale="ru-RU",
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-first-run",
                "--no-default-browser-check",
            ],
        )
        page = context.pages[0] if context.pages else await context.new_page()
        try:
            stagnant_pages = 0
            for page_no in range(0, max_pages):
                if len(vacancies) >= max_results:
                    break
                url = build_search_url(query, area, page_no)
                log(f"Страница выдачи {page_no + 1}: {url}")
                try:
                    await page.goto(url, wait_until="domcontentloaded", timeout=70_000)
                except PlaywrightTimeoutError:
                    log("  ! таймаут страницы выдачи, пропускаю")
                    continue
                await page.wait_for_timeout(4000)
                if page_no == 0:
                    await safe_click_text(page, ("Принять", "Хорошо", "Понятно"))
                await wait_if_blocked(page, log)

                raw_items = await extract_serp_items(page)
                new_count = 0
                for raw in raw_items:
                    clean_url = canonical_vacancy_url(raw.get("href") or "")
                    if not clean_url or vacancy_key(clean_url) in seen:
                        continue
                    seen.add(vacancy_key(clean_url))
                    vacancy = HHVacancy(
                        title=clean_text(raw.get("title")),
                        salary=clean_text(raw.get("salary")),
                        employer=clean_text(raw.get("employer")),
                        employer_hh_url=clean_text(raw.get("employerHref")).split("?")[0],
                        area=clean_text(raw.get("area")),
                        url=clean_url,
                        source_query=query,
                        source_city=city,
                        collected_at=time.strftime("%Y-%m-%d %H:%M:%S"),
                    )
                    if not vacancy.title:
                        continue
                    vacancies.append(vacancy)
                    new_count += 1
                    if not deep:
                        if item_callback:
                            item_callback(vacancy)
                        log(f"  + {vacancy.title} | {vacancy.employer or 'без компании'}")
                    if len(vacancies) >= max_results:
                        break

                log(f"Собрано вакансий: {len(vacancies)}")
                if new_count == 0:
                    stagnant_pages += 1
                    if stagnant_pages >= 2:
                        log("Новых вакансий нет, останавливаю листание.")
                        break
                else:
                    stagnant_pages = 0

            if deep:
                log(f"Открываю страницы вакансий: {len(vacancies)}")
                for index, vacancy in enumerate(vacancies, start=1):
                    log(f"[{index}/{len(vacancies)}] {vacancy.url}")
                    await enrich_from_vacancy_page(page, vacancy, log)
                    if item_callback:
                        item_callback(vacancy)
                    await page.wait_for_timeout(1100)
        finally:
            await context.close()

    paths = write_outputs(vacancies, output_dir)
    with_contacts = sum(1 for v in vacancies if v.contact_email or v.contact_phones)
    log(f"Готово. Вакансий: {len(vacancies)}, с открытыми контактами: {with_contacts}")
    log(f"Excel: {paths['xlsx']}")
    return vacancies, paths


def write_outputs(vacancies: list[HHVacancy], output_dir: Path) -> dict[str, Path]:
    stamp = time.strftime("%Y%m%d_%H%M%S")
    csv_path = output_dir / f"hh_vacancies_{stamp}.csv"
    xlsx_path = output_dir / f"hh_vacancies_{stamp}.xlsx"
    json_path = output_dir / f"hh_vacancies_{stamp}.json"
    fields = list(HHVacancy.__dataclass_fields__.keys())

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for vacancy in vacancies:
            writer.writerow(asdict(vacancy))

    json_path.write_text(
        json.dumps([asdict(vacancy) for vacancy in vacancies], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "HH Vacancies"
    ws.append(fields)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")
    for vacancy in vacancies:
        ws.append([getattr(vacancy, field) for field in fields])
    widths = {
        "A": 42,
        "B": 22,
        "C": 32,
        "D": 34,
        "E": 20,
        "F": 20,
        "G": 26,
        "H": 26,
        "I": 60,
        "J": 16,
        "K": 34,
        "L": 18,
        "M": 16,
        "N": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(xlsx_path)
    return {"csv": csv_path, "xlsx": xlsx_path, "json": json_path}


class HHParserApp:
    def __init__(self) -> None:
        self.root = tk.Tk()
        self.root.title(APP_TITLE)
        self.root.geometry("1180x760")
        self.queue: queue.Queue[tuple[str, object]] = queue.Queue()
        self.worker: Optional[threading.Thread] = None
        self.last_output_dir = DEFAULT_OUTPUT_DIR
        self._build_ui()
        self.root.after(150, self._drain_queue)

    def _build_ui(self) -> None:
        self.root.columnconfigure(1, weight=1)
        self.root.rowconfigure(1, weight=1)

        left = ttk.Frame(self.root, padding=12)
        left.grid(row=0, column=0, rowspan=2, sticky="nsew")
        left.columnconfigure(0, weight=1)

        title = ttk.Label(left, text="HH.ru Parser", font=("Segoe UI", 16, "bold"))
        title.grid(row=0, column=0, sticky="w", pady=(0, 12))

        self.query_var = tk.StringVar(value="менеджер по закупкам")
        self.city_var = tk.StringVar(value="Москва")
        self.limit_var = tk.IntVar(value=100)
        self.pages_var = tk.IntVar(value=10)
        self.deep_var = tk.BooleanVar(value=False)
        self.visible_var = tk.BooleanVar(value=True)
        self.output_var = tk.StringVar(value=str(DEFAULT_OUTPUT_DIR))

        ttk.Label(left, text="Что искать (текст вакансии)").grid(row=1, column=0, sticky="w")
        ttk.Entry(left, textvariable=self.query_var, width=34).grid(row=2, column=0, sticky="ew", pady=(2, 10))

        ttk.Label(left, text="Город / регион").grid(row=3, column=0, sticky="w")
        city = ttk.Combobox(left, textvariable=self.city_var, values=CITY_CHOICES, width=32)
        city.grid(row=4, column=0, sticky="ew", pady=(2, 10))

        ttk.Label(left, text="Лимит вакансий").grid(row=5, column=0, sticky="w")
        ttk.Spinbox(left, from_=1, to=2000, textvariable=self.limit_var, width=10).grid(row=6, column=0, sticky="w", pady=(2, 10))

        ttk.Label(left, text="Максимум страниц выдачи").grid(row=7, column=0, sticky="w")
        ttk.Spinbox(left, from_=1, to=40, textvariable=self.pages_var, width=10).grid(row=8, column=0, sticky="w", pady=(2, 10))

        ttk.Checkbutton(
            left,
            text="Открывать каждую вакансию\n(медленнее: контакты и описание)",
            variable=self.deep_var,
        ).grid(row=9, column=0, sticky="w", pady=(0, 6))

        ttk.Checkbutton(left, text="Показывать Chrome во время сбора", variable=self.visible_var).grid(
            row=10, column=0, sticky="w", pady=(0, 10)
        )

        ttk.Label(left, text="Папка результатов").grid(row=11, column=0, sticky="w")
        ttk.Entry(left, textvariable=self.output_var, width=34).grid(row=12, column=0, sticky="ew", pady=(2, 6))
        ttk.Button(left, text="Выбрать папку", command=self._choose_output_dir).grid(row=13, column=0, sticky="ew", pady=(0, 10))

        self.start_button = ttk.Button(left, text="Запустить сбор", command=self._start)
        self.start_button.grid(row=14, column=0, sticky="ew", pady=(4, 8))
        ttk.Button(left, text="Открыть папку результатов", command=self._open_output_dir).grid(row=15, column=0, sticky="ew")

        hint = ttk.Label(
            left,
            text="Программа открывает обычный Google Chrome\nи листает выдачу hh.ru. Контакты видны\nтолько там, где работодатель их открыл;\nчаще они доступны после входа в аккаунт\nhh.ru в этом же окне Chrome.",
            foreground="#335b9f",
            justify="left",
        )
        hint.grid(row=16, column=0, sticky="w", pady=(18, 0))

        right = ttk.Frame(self.root, padding=(0, 12, 12, 12))
        right.grid(row=0, column=1, rowspan=2, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=3)
        right.rowconfigure(1, weight=2)

        columns = ("title", "employer", "salary", "contacts", "area")
        self.table = ttk.Treeview(right, columns=columns, show="headings", height=16)
        for column, text, width in (
            ("title", "Вакансия", 260),
            ("employer", "Компания", 200),
            ("salary", "Зарплата", 130),
            ("contacts", "Контакты", 220),
            ("area", "Локация", 120),
        ):
            self.table.heading(column, text=text)
            self.table.column(column, width=width, anchor="w")
        self.table.grid(row=0, column=0, sticky="nsew")
        table_scroll = ttk.Scrollbar(right, orient="vertical", command=self.table.yview)
        table_scroll.grid(row=0, column=1, sticky="ns")
        self.table.configure(yscrollcommand=table_scroll.set)

        log_frame = ttk.LabelFrame(right, text="Лог")
        log_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(10, 0))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        self.log_text = tk.Text(log_frame, height=10, bg="#101827", fg="#e6eefb", insertbackground="#e6eefb")
        self.log_text.grid(row=0, column=0, sticky="nsew")
        log_scroll = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        log_scroll.grid(row=0, column=1, sticky="ns")
        self.log_text.configure(yscrollcommand=log_scroll.set)

        self.status_var = tk.StringVar(value="Готов к работе")
        status = ttk.Label(self.root, textvariable=self.status_var, padding=(12, 6), foreground="#1b4f9c")
        status.grid(row=2, column=0, columnspan=2, sticky="ew")

    def _choose_output_dir(self) -> None:
        selected = filedialog.askdirectory(initialdir=self.output_var.get() or str(DEFAULT_OUTPUT_DIR))
        if selected:
            self.output_var.set(selected)

    def _open_output_dir(self) -> None:
        path = Path(self.output_var.get() or self.last_output_dir)
        path.mkdir(parents=True, exist_ok=True)
        subprocess.Popen(["explorer", str(path)])

    def _start(self) -> None:
        if self.worker and self.worker.is_alive():
            messagebox.showinfo(APP_TITLE, "Сбор уже идет.")
            return
        query = clean_text(self.query_var.get())
        city = clean_text(self.city_var.get())
        if not query:
            messagebox.showwarning(APP_TITLE, "Введите запрос: например, менеджер по закупкам.")
            return
        max_results = max(1, int(self.limit_var.get() or 1))
        max_pages = max(1, int(self.pages_var.get() or 1))
        output_dir = Path(self.output_var.get() or DEFAULT_OUTPUT_DIR)
        self.last_output_dir = output_dir
        for row in self.table.get_children():
            self.table.delete(row)
        self.log_text.delete("1.0", "end")
        self.start_button.configure(state="disabled")
        self.status_var.set("Сбор запущен")

        def worker() -> None:
            try:
                asyncio.run(
                    scrape_hh(
                        query=query,
                        city=city,
                        max_results=max_results,
                        max_pages=max_pages,
                        deep=bool(self.deep_var.get()),
                        visible=bool(self.visible_var.get()),
                        output_dir=output_dir,
                        log=lambda message: self.queue.put(("log", message)),
                        item_callback=lambda item: self.queue.put(("item", item)),
                    )
                )
                self.queue.put(("done", output_dir))
            except Exception as exc:
                self.queue.put(("error", str(exc)))

        self.worker = threading.Thread(target=worker, daemon=True)
        self.worker.start()

    def _drain_queue(self) -> None:
        try:
            while True:
                event, payload = self.queue.get_nowait()
                if event == "log":
                    self._log(str(payload))
                elif event == "item" and isinstance(payload, HHVacancy):
                    contacts = "; ".join(filter(None, (payload.contact_email, payload.contact_phones)))
                    self.table.insert(
                        "",
                        "end",
                        values=(payload.title, payload.employer, payload.salary, contacts, payload.area),
                    )
                elif event == "done":
                    self.start_button.configure(state="normal")
                    self.status_var.set("Готово")
                    self._log("Сбор завершен.")
                    messagebox.showinfo(APP_TITLE, "Сбор завершен. Excel и CSV сохранены в папку результатов.")
                elif event == "error":
                    self.start_button.configure(state="normal")
                    self.status_var.set("Ошибка")
                    self._log(f"Ошибка: {payload}")
                    messagebox.showerror(APP_TITLE, str(payload))
        except queue.Empty:
            pass
        self.root.after(150, self._drain_queue)

    def _log(self, message: str) -> None:
        self.log_text.insert("end", f"{time.strftime('%H:%M:%S')} | {message}\n")
        self.log_text.see("end")

    def run(self) -> None:
        self.root.mainloop()


def parse_args(argv: Optional[list[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Desktop HH.ru vacancy parser")
    parser.add_argument("--cli", action="store_true", help="Run without desktop UI")
    parser.add_argument("--query", default="менеджер по закупкам")
    parser.add_argument("--city", default="Москва")
    parser.add_argument("--limit", type=int, default=100)
    parser.add_argument("--pages", type=int, default=10)
    parser.add_argument("--deep", action="store_true", help="Open every vacancy page for contacts")
    parser.add_argument("--headless", action="store_true")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args(argv)


def main(argv: Optional[list[str]] = None) -> int:
    args = parse_args(argv)
    if args.cli:
        for stream in (sys.stdout, sys.stderr):
            try:
                stream.reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass
        vacancies, paths = asyncio.run(
            scrape_hh(
                query=args.query,
                city=args.city,
                max_results=args.limit,
                max_pages=args.pages,
                deep=args.deep,
                visible=not args.headless,
                output_dir=args.output_dir,
                log=print,
            )
        )
        print(f"Saved {len(vacancies)} vacancies")
        print(paths["xlsx"])
        return 0
    app = HHParserApp()
    app.run()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

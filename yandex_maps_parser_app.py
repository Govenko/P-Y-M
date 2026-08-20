#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Standalone desktop Yandex Maps lead parser that drives Google Chrome."""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import os
import queue
import re
import subprocess
import sys
import threading
import time
import urllib.parse
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Callable, Iterable, Optional

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from playwright.async_api import TimeoutError as PlaywrightTimeoutError
from playwright.async_api import async_playwright


APP_TITLE = "Yandex Maps Lead Parser"
ROOT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = ROOT_DIR / "results" / "yandex_maps"
DEFAULT_PROFILE_DIR = ROOT_DIR / "tools" / "yandex_maps_chrome_profile"

PHONE_RE = re.compile(
    r"(?:\+7|8)\s*[\( -]?\d{3,5}[\) -]?\s*\d{1,3}[\s-]?\d{2,3}[\s-]?\d{2,3}"
)
ORG_URL_RE = re.compile(r"(?P<prefix>https?://[^/]+)?/maps/org/(?P<slug>[^/?#]+)/(?P<id>\d+)/?")
BAD_SITE_DOMAINS = (
    "yandex.",
    "ya.ru",
    "google.",
    "2gis.",
    "vk.com",
    "t.me",
    "telegram.",
    "wa.me",
    "whatsapp.",
    "max.ru",
    "instagram.",
    "youtube.",
)
SOCIAL_DOMAINS = ("vk.com", "t.me", "telegram.", "wa.me", "whatsapp.", "max.ru", "instagram.", "youtube.")

CITY_PATHS = {
    "вся россия": "225/russia",
    "россия": "225/russia",
    "москва": "213/moscow",
    "санкт-петербург": "2/saint-petersburg",
    "спб": "2/saint-petersburg",
    "казань": "43/kazan",
    "нижний новгород": "47/nizhny-novgorod",
    "самара": "51/samara",
    "уфа": "172/ufa",
    "екатеринбург": "54/yekaterinburg",
    "пермь": "50/perm",
    "челябинск": "56/chelyabinsk",
    "ростов-на-дону": "39/rostov-na-donu",
    "краснодар": "35/krasnodar",
    "воронеж": "193/voronezh",
    "новосибирск": "65/novosibirsk",
    "тюмень": "55/tyumen",
    "омск": "66/omsk",
    "красноярск": "62/krasnoyarsk",
    "владивосток": "75/vladivostok",
}

CITY_CHOICES = ["Вся Россия"] + [
    name.title() for name in CITY_PATHS if name not in ("вся россия", "россия", "спб")
]

# Режим «Вся Россия»: поиск на Яндекс.Картах работает по видимой области карты,
# поэтому всю страну обходим по крупным городам, дубли убираются по id организации.
RF_SWEEP_CITIES = [
    "Москва", "Санкт-Петербург", "Новосибирск", "Екатеринбург", "Казань",
    "Нижний Новгород", "Челябинск", "Самара", "Омск", "Ростов-на-Дону",
    "Уфа", "Красноярск", "Воронеж", "Пермь", "Волгоград", "Краснодар",
    "Саратов", "Тюмень", "Тольятти", "Ижевск", "Барнаул", "Ульяновск",
    "Иркутск", "Хабаровск", "Ярославль", "Владивосток", "Махачкала",
    "Томск", "Оренбург", "Кемерово", "Новокузнецк", "Рязань", "Астрахань",
    "Набережные Челны", "Пенза", "Липецк", "Киров", "Чебоксары", "Тула",
    "Калининград", "Курск", "Ставрополь", "Сочи", "Улан-Удэ", "Тверь",
    "Магнитогорск", "Иваново", "Брянск", "Белгород", "Сургут", "Владимир",
    "Архангельск", "Чита", "Симферополь", "Севастополь", "Калуга", "Смоленск",
    "Волжский", "Мурманск", "Владикавказ", "Грозный", "Тамбов", "Вологда",
    "Якутск", "Петрозаводск",
]


def is_all_russia(city: str) -> bool:
    return clean_text(city).lower().replace("ё", "е") in ("", "вся россия", "россия")


@dataclass
class YandexLead:
    name: str = ""
    category: str = ""
    address: str = ""
    phones: str = ""
    website: str = ""
    telegram: str = ""
    whatsapp: str = ""
    vk: str = ""
    instagram: str = ""
    youtube: str = ""
    other_socials: str = ""
    rating: str = ""
    reviews: str = ""
    working_status: str = ""
    working_hours: str = ""
    logo: str = ""
    images: str = ""
    yandex_url: str = ""
    source_query: str = ""
    source_city: str = ""
    collected_at: str = ""
    raw_text: str = ""


def clean_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def unique(values: Iterable[str], limit: int = 20) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        clean = clean_text(value).strip(" ,;")
        key = clean.lower()
        if clean and key not in seen:
            seen.add(key)
            result.append(clean)
            if len(result) >= limit:
                break
    return result


def find_chrome() -> Path:
    candidates = [
        Path(os.environ.get("ProgramFiles", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
        Path(os.environ.get("ProgramFiles(x86)", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
        Path(os.environ.get("LOCALAPPDATA", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
        Path("/usr/bin/google-chrome"),
        Path("/usr/bin/google-chrome-stable"),
        # Playwright Chromium fallback
        Path("/root/.cache/ms-playwright/chromium-1117/chrome-linux/chrome"),
        Path.home() / ".cache" / "ms-playwright" / "chromium-1117" / "chrome-linux" / "chrome",
    ]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError("Google Chrome не найден. Установите Chrome и запустите программу снова.")


def normalize_phone(value: str) -> str:
    match = PHONE_RE.search(value)
    if match:
        value = match.group(0)
    digits = re.sub(r"\D", "", value)
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    if len(digits) == 10:
        digits = "7" + digits
    if len(digits) == 11 and digits.startswith("7"):
        return f"+7 ({digits[1:4]}) {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
    return clean_text(value)


def canonical_org_url(url: str) -> str:
    match = ORG_URL_RE.search(url or "")
    if not match:
        return ""
    return f"https://yandex.ru/maps/org/{match.group('slug')}/{match.group('id')}/"


def org_key(url: str) -> str:
    match = ORG_URL_RE.search(url or "")
    return match.group("id") if match else url


def domain_of(url: str) -> str:
    try:
        return urllib.parse.urlsplit(url).netloc.lower().removeprefix("www.")
    except Exception:
        return ""


def build_search_url(query: str, city: str) -> str:
    city_key = city.lower().replace("ё", "е").strip()
    if city_key in ("", "вся россия", "россия"):
        # По всей РФ город к запросу не добавляем, ищем в регионе "Россия".
        encoded = urllib.parse.quote(clean_text(query))
        return f"https://yandex.ru/maps/225/russia/search/{encoded}/"
    full_query = clean_text(f"{query} {city}".strip())
    encoded = urllib.parse.quote(full_query)
    city_path = CITY_PATHS.get(city_key)
    if city_path:
        return f"https://yandex.ru/maps/{city_path}/search/{encoded}/"
    return f"https://yandex.ru/maps/?text={encoded}"


async def safe_click_text(page, texts: Iterable[str], timeout: int = 1200) -> None:
    for text in texts:
        try:
            locator = page.get_by_text(text, exact=False).first
            if await locator.count():
                await locator.click(timeout=timeout)
                await page.wait_for_timeout(400)
                return
        except Exception:
            continue


async def accept_cookies(page) -> None:
    await safe_click_text(
        page,
        (
            "Allow all",
            "Accept all",
            "Принять все",
            "Разрешить все",
            "Хорошо",
            "Понятно",
            "Согласен",
        ),
    )


async def find_scroll_container(page):
    return await page.evaluate_handle(
        """
        () => {
          const preferred = document.querySelector('.scroll__container')
            || document.querySelector('.search-list-view')
            || document.querySelector('[class*="search-list"]');
          if (preferred && preferred.scrollHeight > preferred.clientHeight) return preferred;
          const candidates = Array.from(document.querySelectorAll('div,section,ul'))
            .filter(e => e.scrollHeight > e.clientHeight + 200)
            .sort((a, b) => b.scrollHeight - a.scrollHeight);
          return candidates[0] || document.scrollingElement || document.body;
        }
        """
    )


async def collect_org_links(page, max_results: int, log: Callable[[str], None]) -> list[str]:
    links: list[str] = []
    seen: set[str] = set()
    stagnant_rounds = 0
    last_count = 0

    max_rounds = max(80, min(500, max_results * 3))
    for round_no in range(1, max_rounds):
        # Пробуем разные селекторы для поиска карточек организаций
        found = await page.evaluate(
            """
            () => {
              const selectors = [
                'a[href*="/maps/org/"]',
                'a[href*="/org/"]',
                '.business-card-view a',
                '.search-list-item a',
                '[data-id] a[href*="/maps/"]',
                '.ymaps-bizcard a'
              ];
              const results = [];
              selectors.forEach(sel => {
                document.querySelectorAll(sel).forEach(a => {
                  const href = a.href || '';
                  if (href) results.push(href);
                });
              });
              return results;
            }
            """
        )
        for href in found:
            clean = canonical_org_url(href)
            if clean and org_key(clean) not in seen:
                seen.add(org_key(clean))
                links.append(clean)
                if len(links) >= max_results:
                    return links

        if len(links) == last_count:
            stagnant_rounds += 1
        else:
            stagnant_rounds = 0
            last_count = len(links)
            log(f"Найдено карточек: {len(links)}")

        if stagnant_rounds >= 10 and links:
            break

        scroll = await find_scroll_container(page)
        await scroll.evaluate(
            """
            element => {
              const distance = Math.max(900, element.clientHeight || 900);
              element.scrollTop = (element.scrollTop || 0) + distance;
              window.dispatchEvent(new Event('scroll'));
            }
            """
        )
        await page.wait_for_timeout(900 + min(round_no, 8) * 80)

    return links[:max_results]


async def extract_current_org(page, query: str, city: str) -> YandexLead:
    await safe_click_text(page, ("Показать телефон", "Показать номер", "Show phone"), timeout=900)
    await page.wait_for_timeout(700)
    data = await page.evaluate(
        """
        () => {
          const pick = selectors => {
            for (const selector of selectors) {
              const el = document.querySelector(selector);
              const text = el && (el.innerText || el.textContent || '').trim();
              if (text) return text;
            }
            return '';
          };
          const allText = document.body.innerText || '';
          const links = Array.from(document.querySelectorAll('a')).map(a => ({
            text: (a.innerText || a.textContent || '').trim(),
            href: a.href || ''
          })).filter(x => x.href);
          const categoryLinks = Array.from(document.querySelectorAll('a[href*="/category/"], button'))
            .map(e => (e.innerText || e.textContent || '').trim())
            .filter(Boolean)
            .slice(0, 8);
          
          // Извлекаем полный график работы по дням недели
          let workingHours = '';
          const scheduleRows = document.querySelectorAll(
            '[class*="schedule"] tr, ' +
            '[class*="working-hours"] tr, ' +
            '.business-schedule-view__row, ' +
            '.business-schedule-view__table tr, ' +
            '[class*="schedule-table"] tr, ' +
            '.orgpage-schedule-view__row'
          );
          if (scheduleRows.length > 0) {
            const hoursParts = [];
            scheduleRows.forEach(row => {
              const text = row.innerText || row.textContent || '';
              if (text.trim()) hoursParts.push(text.trim());
            });
            workingHours = hoursParts.join('; ');
          } else {
            // Альтернативные селекторы для графика - блочный формат
            const scheduleBlock = pick([
              '.business-schedule-view',
              '[class*="working-hours"]',
              '[class*="schedule-block"]',
              '.orgpage-schedule-view',
              '[class*="hours-list"]'
            ]);
            if (scheduleBlock) workingHours = scheduleBlock;
          }
          
          // Если график не найден, пробуем найти по отдельным элементам дней
          if (!workingHours) {
            const days = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс', 'понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота', 'воскресенье'];
            const dayElements = document.querySelectorAll('[class*="day"], [class*="weekday"]');
            if (dayElements.length > 0) {
              const hoursParts = [];
              dayElements.forEach(el => {
                const text = el.innerText || el.textContent || '';
                if (text.trim() && days.some(d => text.toLowerCase().includes(d))) {
                  hoursParts.push(text.trim());
                }
              });
              if (hoursParts.length > 0) workingHours = hoursParts.join('; ');
            }
          }
          
          // Извлекаем изображения (логотип и фото организации)
          const images = [];
          const logoImages = document.querySelectorAll(
            '.business-card-logo-view__image img, ' +
            '.orgpage-header-view__logo img, ' +
            '.business-header-view__logo img, ' +
            '[class*="logo"] img, ' +
            'img[class*="logo"]'
          );
          logoImages.forEach(img => {
            const src = img.dataset.imageUrl || img.src || '';
            if (src && !images.includes(src)) images.push(src);
          });
          
          // Собираем все изображения из галереи
          const galleryImages = document.querySelectorAll(
            '.business-gallery-view__item img, ' +
            '[class*="gallery"] img, ' +
            '.orgpage-photos-view__photo img, ' +
            '.business-photos-view__photo img, ' +
            '[data-image-url], ' +
            'img[src*="/maps/"], ' +
            '.business-card-carousel__image img, ' +
            '[class*="carousel"] img, ' +
            '[class*="photo-item"] img'
          );
          galleryImages.forEach(img => {
            const src = img.dataset.imageUrl || img.src || '';
            if (src && !images.includes(src)) images.push(src);
          });
          
          // Также пробуем найти изображения через data-атрибуты
          const dataImages = document.querySelectorAll('[data-image-url], [data-src], [data-lazy-src]');
          dataImages.forEach(el => {
            const src = el.dataset.imageUrl || el.dataset.src || el.dataset.lazySrc || '';
            if (src && !images.includes(src)) images.push(src);
          });
          
          return {
            title: document.title || '',
            pageUrl: location.href,
            name: pick([
              '.orgpage-header-view__title',
              '.business-card-title-view__title',
              '.business-header-view__title',
              'h1'
            ]),
            breadcrumbs: pick(['.business-card-view__breadcrumbs']),
            address: pick([
              '.orgpage-header-view__address',
              '.business-contacts-view__address',
              'a[href*="/house/"]',
              '.business-address-view'
            ]),
            rating: pick([
              '.business-rating-badge-view__rating-text',
              '.business-rating-badge-view__rating',
              '.business-header-rating-view__rating'
            ]),
            reviews: pick([
              '.business-header-rating-view__text',
              'a[href$="/reviews/"]',
              '[class*="reviews-count"]'
            ]),
            workingStatus: pick([
              '.business-working-status-view',
              '.business-working-status-flip-view',
              '[class*="working-status"]'
            ]),
            workingHours: workingHours,
            phoneTexts: Array.from(document.querySelectorAll(
              '.orgpage-phones-view__phone-number, ' +
              '[class*="phone-number"], ' +
              '[class*="phone"] a[href^="tel:"]'
            ))
              .map(e => (e.innerText || e.textContent || '').trim())
              .filter(Boolean),
            categoryLinks,
            links,
            images,
            allText: allText.slice(0, 5000)
          };
        }
        """
    )

    title = clean_text(data.get("title"))
    name = clean_text(data.get("name")) or clean_text(title.split(",")[0])
    breadcrumbs = clean_text(data.get("breadcrumbs"))
    categories = [part.strip() for part in re.split(r"[·•>]", breadcrumbs) if part.strip()]
    # Убираем мусор из категорий: "Маршрут", "Исправить неточность" и т.д.
    trash_words = ["Маршрут", "Исправить неточность", "Поделиться", "Написать отзыв"]
    category_candidates = []
    for cat in categories[2:] if len(categories) >= 3 else categories:
        is_trash = False
        for trash in trash_words:
            if trash in cat:
                is_trash = True
                break
        if not is_trash and cat:
            category_candidates.append(cat)
    
    category_candidates.extend(data.get("categoryLinks") or [])
    # Также фильтруем categoryLinks от мусора
    cleaned_links = []
    for link in category_candidates:
        is_trash = False
        for trash in trash_words:
            if trash in link:
                is_trash = True
                break
        if not is_trash and link:
            cleaned_links.append(link)
    category = "; ".join(unique(cleaned_links, limit=5))

    all_text = str(data.get("allText") or "")
    phone_candidates = list(data.get("phoneTexts") or [])
    phone_candidates.extend(PHONE_RE.findall(all_text))
    phones = "; ".join(unique(normalize_phone(phone) for phone in phone_candidates if phone))

    websites: list[str] = []
    telegram_links: list[str] = []
    whatsapp_links: list[str] = []
    vk_links: list[str] = []
    instagram_links: list[str] = []
    youtube_links: list[str] = []
    other_socials: list[str] = []
    
    # Ссылки на Яндекс которые нужно игнорировать
    yandex_domains = ["yandex.ru", "yandex.com", "ya.ru", "mapsyandex", "yandex.maps"]
    
    for link in data.get("links") or []:
        href = clean_text(link.get("href"))
        if not href.startswith(("http://", "https://")):
            continue
        domain = domain_of(href).lower()
        
        # Пропускаем ссылки на сам Яндекс и технические страницы
        if any(ya in domain for ya in yandex_domains):
            continue
        
        # Распределяем ссылки по категориям
        if "t.me" in domain or "telegram.org" in domain or "telegram.me" in domain:
            telegram_links.append(href)
        elif "wa.me" in domain or "whatsapp.com" in domain or "api.whatsapp.com" in domain:
            whatsapp_links.append(href)
        elif "vk.com" in domain or "vk.ru" in domain:
            vk_links.append(href)
        elif "instagram.com" in domain or "instagr.am" in domain:
            instagram_links.append(href)
        elif "youtube.com" in domain or "youtu.be" in domain:
            youtube_links.append(href)
        elif any(part in domain for part in SOCIAL_DOMAINS):
            other_socials.append(href)

    # Фильтруем изображения от JS-скриптов и мусора
    def is_valid_image_url(url):
        if not url:
            return False
        # Отсекаем технические скрипты и CSS
        if '.js' in url or '.css' in url:
            return False
        if 'yastatic' in url and 'maps-front' in url:
            return False
        # Оставляем только картинки
        image_extensions = ['.jpg', '.jpeg', '.png', '.webp', '.gif', '.svg']
        if any(ext in url.lower() for ext in image_extensions):
            return True
        # Если ссылка содержит /i/ или /images/ и не скрипт
        if '/i/' in url or '/images/' in url or 'avatars' in url or 'XXL_height' in url:
            return True
        return False
    
    # Извлекаем изображения из карточки организации
    images_list = data.get("images") or []
    valid_images = [img for img in images_list if is_valid_image_url(img)]
    
    # Отделяем логотип от остальных изображений
    logo_url = ""
    other_images = []
    for img in valid_images:
        if 'logo' in img.lower():
            # Это точно логотип
            if not logo_url:
                logo_url = img
            else:
                other_images.append(img)
        elif 'XXL_height' in img or 'sugoi' in img or 'avatars' in img:
            # Это фото организации
            other_images.append(img)
        else:
            # Неизвестное изображение, добавляем в другие
            other_images.append(img)
    
    # Если не нашли логотип, берем первое изображение (если оно не мусор)
    if not logo_url and valid_images:
        logo_url = valid_images[0]
        other_images = valid_images[1:]
    
    return YandexLead(
        name=name,
        category=category,
        address=clean_text(data.get("address")),
        phones=phones,
        website="; ".join(unique(websites, limit=5)),
        telegram="; ".join(unique(telegram_links, limit=3)),
        whatsapp="; ".join(unique(whatsapp_links, limit=3)),
        vk="; ".join(unique(vk_links, limit=3)),
        instagram="; ".join(unique(instagram_links, limit=3)),
        youtube="; ".join(unique(youtube_links, limit=3)),
        other_socials="; ".join(unique(other_socials, limit=5)),
        rating=clean_text(data.get("rating")).replace("Рейтинг ", ""),
        reviews=clean_text(data.get("reviews")),
        working_status=clean_text(data.get("workingStatus")),
        working_hours=clean_text(data.get("workingHours")),
        logo=logo_url,
        images="; ".join(unique(other_images, limit=20)),
        yandex_url=canonical_org_url(data.get("pageUrl") or page.url) or page.url,
        source_query=query,
        source_city=city,
        collected_at=time.strftime("%Y-%m-%d %H:%M:%S"),
        raw_text=clean_text(all_text)[:900],
    )


async def scrape_yandex_maps(
    *,
    query: str,
    city: str,
    max_results: int,
    visible: bool,
    output_dir: Path,
    log: Callable[[str], None],
    lead_callback: Optional[Callable[[YandexLead], None]] = None,
) -> tuple[list[YandexLead], dict[str, Path]]:
    chrome = find_chrome()
    DEFAULT_PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    if is_all_russia(city):
        sweep_cities = RF_SWEEP_CITIES
    else:
        sweep_cities = [city]

    log(f"Chrome: {chrome}")
    if len(sweep_cities) > 1:
        log(f"Режим «Вся Россия»: обход {len(sweep_cities)} городов, пока не наберется {max_results} организаций.")
    else:
        log(f"Открываю Яндекс.Карты: {query} / {city}")
    leads: list[YandexLead] = []
    seen_orgs: set[str] = set()

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            str(DEFAULT_PROFILE_DIR),
            executable_path=str(chrome),
            headless=not visible,
            viewport={"width": 1360, "height": 900},
            locale="ru-RU",
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-first-run",
                "--no-default-browser-check",
            ],
        )
        page = context.pages[0] if context.pages else await context.new_page()
        try:
            first_page = True
            for city_index, sweep_city in enumerate(sweep_cities, start=1):
                if len(leads) >= max_results:
                    break
                if len(sweep_cities) > 1:
                    log(f"— Город {city_index}/{len(sweep_cities)}: {sweep_city} (собрано {len(leads)}/{max_results})")
                start_url = build_search_url(query, sweep_city)
                try:
                    await page.goto(start_url, wait_until="domcontentloaded", timeout=70_000)
                except PlaywrightTimeoutError:
                    log("  ! таймаут поиска, пропускаю город")
                    continue
                await page.wait_for_timeout(7000 if first_page else 4000)
                if first_page:
                    await accept_cookies(page)
                    log(f"Страница открыта: {await page.title()}")
                    first_page = False

                if "captcha" in page.url.lower() or "showcaptcha" in (await page.content()).lower():
                    log("Яндекс показал проверку. Решите ее в открытом Chrome, программа подождет 90 секунд.")
                    await page.wait_for_timeout(90_000)

                # Даем время на загрузку результатов поиска
                await page.wait_for_timeout(5000)
                
                remaining = max_results - len(leads)
                org_links = await collect_org_links(page, max_results=remaining, log=log)
                # Если запрос сразу открыл карточку организации, берем и ее.
                direct = canonical_org_url(page.url)
                if direct and direct not in org_links:
                    org_links.insert(0, direct)
                org_links = [url for url in org_links if org_key(url) not in seen_orgs]
                log(f"Буду обрабатывать карточек: {len(org_links)}")

                for index, url in enumerate(org_links, start=1):
                    if len(leads) >= max_results:
                        break
                    seen_orgs.add(org_key(url))
                    try:
                        log(f"[{index}/{len(org_links)}] {url}")
                        await page.goto(url, wait_until="domcontentloaded", timeout=60_000)
                        await page.wait_for_timeout(2800)
                        lead = await extract_current_org(page, query=query, city=sweep_city)
                        if lead.name:
                            leads.append(lead)
                            if lead_callback:
                                lead_callback(lead)
                            log(f"  + {lead.name} | {lead.phones or 'нет телефона'} | {lead.website or 'нет сайта'}")
                    except PlaywrightTimeoutError:
                        log("  ! таймаут карточки, пропускаю")
                    except Exception as exc:
                        log(f"  ! ошибка карточки: {exc}")
        finally:
            await context.close()

    leads = dedupe_leads(leads)
    paths = write_outputs(leads, output_dir)
    log(f"Готово. Лидов: {len(leads)}")
    log(f"Excel: {paths['xlsx']}")
    return leads, paths


def dedupe_leads(leads: Iterable[YandexLead]) -> list[YandexLead]:
    result: list[YandexLead] = []
    seen: set[str] = set()
    for lead in leads:
        key = org_key(lead.yandex_url) or lead.website or (lead.name + lead.address).lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(lead)
    return result


def write_outputs(leads: list[YandexLead], output_dir: Path) -> dict[str, Path]:
    stamp = time.strftime("%Y%m%d_%H%M%S")
    csv_path = output_dir / f"yandex_maps_leads_{stamp}.csv"
    xlsx_path = output_dir / f"yandex_maps_leads_{stamp}.xlsx"
    json_path = output_dir / f"yandex_maps_leads_{stamp}.json"
    fields = list(YandexLead.__dataclass_fields__.keys())

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for lead in leads:
            writer.writerow(asdict(lead))

    json_path.write_text(json.dumps([asdict(lead) for lead in leads], ensure_ascii=False, indent=2), encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "Yandex Maps Leads"
    ws.append(fields)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")
    for lead in leads:
        ws.append([getattr(lead, field) for field in fields])
    widths = {
        "A": 32,  # name
        "B": 24,  # category
        "C": 42,  # address
        "D": 24,  # phones
        "E": 30,  # website
        "F": 28,  # telegram
        "G": 28,  # whatsapp
        "H": 28,  # vk
        "I": 28,  # instagram
        "J": 28,  # youtube
        "K": 28,  # other_socials
        "L": 10,  # rating
        "M": 18,  # reviews
        "N": 20,  # working_status
        "O": 45,  # working_hours
        "P": 60,  # images
        "Q": 46,  # yandex_url
        "R": 24,  # source_query
        "S": 18,  # source_city
        "T": 20,  # collected_at
        "U": 60,  # raw_text
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(xlsx_path)
    return {"csv": csv_path, "xlsx": xlsx_path, "json": json_path}


class YandexMapsParserApp:
    def __init__(self) -> None:
        self.root = tk.Tk()
        self.root.title(APP_TITLE)
        self.root.geometry("1180x760")
        self.queue: queue.Queue[tuple[str, object]] = queue.Queue()
        self.worker: Optional[threading.Thread] = None
        self.last_output_dir = DEFAULT_OUTPUT_DIR
        self._build_ui()
        self.root.after(150, self._drain_queue)

    def _build_ui(self) -> None:
        self.root.columnconfigure(1, weight=1)
        self.root.rowconfigure(1, weight=1)

        left = ttk.Frame(self.root, padding=12)
        left.grid(row=0, column=0, rowspan=2, sticky="nsew")
        left.columnconfigure(0, weight=1)

        title = ttk.Label(left, text="Yandex Maps Lead Parser", font=("Segoe UI", 16, "bold"))
        title.grid(row=0, column=0, sticky="w", pady=(0, 12))

        self.query_var = tk.StringVar(value="кофейня")
        self.city_var = tk.StringVar(value="Москва")
        self.limit_var = tk.IntVar(value=80)
        self.visible_var = tk.BooleanVar(value=True)
        self.output_var = tk.StringVar(value=str(DEFAULT_OUTPUT_DIR))

        ttk.Label(left, text="Что искать").grid(row=1, column=0, sticky="w")
        ttk.Entry(left, textvariable=self.query_var, width=34).grid(row=2, column=0, sticky="ew", pady=(2, 10))

        ttk.Label(left, text="Город").grid(row=3, column=0, sticky="w")
        city = ttk.Combobox(left, textvariable=self.city_var, values=CITY_CHOICES, width=32)
        city.grid(row=4, column=0, sticky="ew", pady=(2, 10))

        ttk.Label(left, text="Лимит организаций").grid(row=5, column=0, sticky="w")
        ttk.Spinbox(left, from_=1, to=1000, textvariable=self.limit_var, width=10).grid(row=6, column=0, sticky="w", pady=(2, 10))

        ttk.Checkbutton(left, text="Показывать Chrome во время сбора", variable=self.visible_var).grid(
            row=7, column=0, sticky="w", pady=(0, 10)
        )

        ttk.Label(left, text="Папка результатов").grid(row=8, column=0, sticky="w")
        ttk.Entry(left, textvariable=self.output_var, width=34).grid(row=9, column=0, sticky="ew", pady=(2, 6))
        ttk.Button(left, text="Выбрать папку", command=self._choose_output_dir).grid(row=10, column=0, sticky="ew", pady=(0, 10))

        self.start_button = ttk.Button(left, text="Запустить сбор", command=self._start)
        self.start_button.grid(row=11, column=0, sticky="ew", pady=(4, 8))
        ttk.Button(left, text="Открыть папку результатов", command=self._open_output_dir).grid(row=12, column=0, sticky="ew")

        hint = ttk.Label(
            left,
            text="Программа открывает обычный Google Chrome,\nлистает Яндекс.Карты и сохраняет Excel.",
            foreground="#335b9f",
            justify="left",
        )
        hint.grid(row=13, column=0, sticky="w", pady=(18, 0))

        right = ttk.Frame(self.root, padding=(0, 12, 12, 12))
        right.grid(row=0, column=1, rowspan=2, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=3)
        right.rowconfigure(1, weight=2)

        columns = ("name", "phone", "site", "rating", "address")
        self.table = ttk.Treeview(right, columns=columns, show="headings", height=16)
        for column, text, width in (
            ("name", "Компания", 220),
            ("phone", "Телефон", 170),
            ("site", "Сайт", 210),
            ("rating", "Рейтинг", 80),
            ("address", "Адрес", 360),
        ):
            self.table.heading(column, text=text)
            self.table.column(column, width=width, anchor="w")
        self.table.grid(row=0, column=0, sticky="nsew")
        table_scroll = ttk.Scrollbar(right, orient="vertical", command=self.table.yview)
        table_scroll.grid(row=0, column=1, sticky="ns")
        self.table.configure(yscrollcommand=table_scroll.set)

        log_frame = ttk.LabelFrame(right, text="Лог")
        log_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(10, 0))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        self.log_text = tk.Text(log_frame, height=10, bg="#101827", fg="#e6eefb", insertbackground="#e6eefb")
        self.log_text.grid(row=0, column=0, sticky="nsew")
        log_scroll = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        log_scroll.grid(row=0, column=1, sticky="ns")
        self.log_text.configure(yscrollcommand=log_scroll.set)

        self.status_var = tk.StringVar(value="Готов к работе")
        status = ttk.Label(self.root, textvariable=self.status_var, padding=(12, 6), foreground="#1b4f9c")
        status.grid(row=2, column=0, columnspan=2, sticky="ew")

    def _choose_output_dir(self) -> None:
        selected = filedialog.askdirectory(initialdir=self.output_var.get() or str(DEFAULT_OUTPUT_DIR))
        if selected:
            self.output_var.set(selected)

    def _open_output_dir(self) -> None:
        path = Path(self.output_var.get() or self.last_output_dir)
        path.mkdir(parents=True, exist_ok=True)
        subprocess.Popen(["explorer", str(path)])

    def _start(self) -> None:
        if self.worker and self.worker.is_alive():
            messagebox.showinfo(APP_TITLE, "Сбор уже идет.")
            return
        query = clean_text(self.query_var.get())
        city = clean_text(self.city_var.get())
        if not query:
            messagebox.showwarning(APP_TITLE, "Введите запрос: например, кофейня или магазин здорового питания.")
            return
        max_results = max(1, int(self.limit_var.get() or 1))
        output_dir = Path(self.output_var.get() or DEFAULT_OUTPUT_DIR)
        self.last_output_dir = output_dir
        for row in self.table.get_children():
            self.table.delete(row)
        self.log_text.delete("1.0", "end")
        self.start_button.configure(state="disabled")
        self.status_var.set("Сбор запущен")

        def worker() -> None:
            try:
                asyncio.run(
                    scrape_yandex_maps(
                        query=query,
                        city=city,
                        max_results=max_results,
                        visible=bool(self.visible_var.get()),
                        output_dir=output_dir,
                        log=lambda message: self.queue.put(("log", message)),
                        lead_callback=lambda lead: self.queue.put(("lead", lead)),
                    )
                )
                self.queue.put(("done", output_dir))
            except Exception as exc:
                self.queue.put(("error", str(exc)))

        self.worker = threading.Thread(target=worker, daemon=True)
        self.worker.start()

    def _drain_queue(self) -> None:
        try:
            while True:
                event, payload = self.queue.get_nowait()
                if event == "log":
                    self._log(str(payload))
                elif event == "lead" and isinstance(payload, YandexLead):
                    self.table.insert(
                        "",
                        "end",
                        values=(payload.name, payload.phones, payload.website, payload.rating, payload.address),
                    )
                elif event == "done":
                    self.start_button.configure(state="normal")
                    self.status_var.set("Готово")
                    self._log("Сбор завершен.")
                    messagebox.showinfo(APP_TITLE, "Сбор завершен. Excel и CSV сохранены в папку результатов.")
                elif event == "error":
                    self.start_button.configure(state="normal")
                    self.status_var.set("Ошибка")
                    self._log(f"Ошибка: {payload}")
                    messagebox.showerror(APP_TITLE, str(payload))
        except queue.Empty:
            pass
        self.root.after(150, self._drain_queue)

    def _log(self, message: str) -> None:
        self.log_text.insert("end", f"{time.strftime('%H:%M:%S')} | {message}\n")
        self.log_text.see("end")

    def run(self) -> None:
        self.root.mainloop()


def parse_args(argv: Optional[list[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Desktop Yandex Maps parser")
    parser.add_argument("--cli", action="store_true", help="Run without desktop UI")
    parser.add_argument("--query", default="кофейня")
    parser.add_argument("--city", default="Москва")
    parser.add_argument("--limit", type=int, default=80)
    parser.add_argument("--headless", action="store_true")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args(argv)


def main(argv: Optional[list[str]] = None) -> int:
    args = parse_args(argv)
    if args.cli:
        for stream in (sys.stdout, sys.stderr):
            try:
                stream.reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass
        leads, paths = asyncio.run(
            scrape_yandex_maps(
                query=args.query,
                city=args.city,
                max_results=args.limit,
                visible=not args.headless,
                output_dir=args.output_dir,
                log=print,
            )
        )
        print(f"Saved {len(leads)} leads")
        print(paths["xlsx"])
        return 0
    app = YandexMapsParserApp()
    app.run()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

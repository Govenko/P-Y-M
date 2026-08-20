#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Standalone desktop Avito parser that drives Google Chrome."""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import os
import queue
import re
import subprocess
import sys
import threading
import time
import urllib.parse
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Callable, Iterable, Optional

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from playwright.async_api import TimeoutError as PlaywrightTimeoutError
from playwright.async_api import async_playwright


APP_TITLE = "Avito Parser"
ROOT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = ROOT_DIR / "results" / "avito"
DEFAULT_PROFILE_DIR = ROOT_DIR / "tools" / "avito_chrome_profile"

ITEM_URL_RE = re.compile(r"https?://(?:www\.|m\.)?avito\.ru(?P<path>/[^\s\"'?#]+_(?P<id>\d{7,}))")
PRICE_RE = re.compile(r"(\d[\d\s ]*)\s*(?:₽|руб)")

CITY_SLUGS = {
    "вся россия": "all",
    "москва": "moskva",
    "московская область": "moskovskaya_oblast",
    "санкт-петербург": "sankt-peterburg",
    "спб": "sankt-peterburg",
    "казань": "kazan",
    "нижний новгород": "nizhniy_novgorod",
    "самара": "samara",
    "уфа": "ufa",
    "екатеринбург": "ekaterinburg",
    "пермь": "perm",
    "челябинск": "chelyabinsk",
    "ростов-на-дону": "rostov-na-donu",
    "краснодар": "krasnodar",
    "воронеж": "voronezh",
    "новосибирск": "novosibirsk",
    "тюмень": "tyumen",
    "омск": "omsk",
    "красноярск": "krasnoyarsk",
    "владивосток": "vladivostok",
    "ижевск": "izhevsk",
}

BLOCK_MARKERS = (
    "доступ ограничен",
    "проблема с ip",
    "подтвердите, что вы не робот",
    "captcha",
    "firewall",
)


@dataclass
class AvitoItem:
    title: str = ""
    price: str = ""
    url: str = ""
    location: str = ""
    seller_name: str = ""
    description: str = ""
    published: str = ""
    source_query: str = ""
    source_city: str = ""
    collected_at: str = ""


def clean_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def unique(values: Iterable[str], limit: int = 20) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        clean = clean_text(value).strip(" ,;")
        key = clean.lower()
        if clean and key not in seen:
            seen.add(key)
            result.append(clean)
            if len(result) >= limit:
                break
    return result


def find_chrome() -> Path:
    candidates = [
        Path(os.environ.get("ProgramFiles", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
        Path(os.environ.get("ProgramFiles(x86)", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
        Path(os.environ.get("LOCALAPPDATA", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
    ]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError("Google Chrome не найден. Установите Chrome и запустите программу снова.")


def canonical_item_url(url: str) -> str:
    match = ITEM_URL_RE.search(url or "")
    if not match:
        return ""
    return f"https://www.avito.ru{match.group('path')}"


def item_key(url: str) -> str:
    match = ITEM_URL_RE.search(url or "")
    return match.group("id") if match else url


def normalize_price(value: str) -> str:
    value = clean_text(value)
    match = PRICE_RE.search(value)
    if match:
        digits = re.sub(r"\D", "", match.group(1))
        if digits:
            return f"{int(digits):,} ₽".replace(",", " ")
    digits = re.sub(r"\D", "", value)
    if digits and digits == value.replace(" ", ""):
        return f"{int(digits):,} ₽".replace(",", " ")
    return value


def build_search_url(query: str, city: str, page_no: int) -> str:
    city_key = city.lower().replace("ё", "е").strip()
    slug = CITY_SLUGS.get(city_key, "all")
    encoded = urllib.parse.quote(clean_text(query))
    url = f"https://www.avito.ru/{slug}?q={encoded}"
    if page_no > 1:
        url += f"&p={page_no}"
    return url


async def safe_click_text(page, texts: Iterable[str], timeout: int = 1200) -> None:
    for text in texts:
        try:
            locator = page.get_by_text(text, exact=False).first
            if await locator.count():
                await locator.click(timeout=timeout)
                await page.wait_for_timeout(400)
                return
        except Exception:
            continue


async def accept_cookies(page) -> None:
    await safe_click_text(page, ("Принять все", "Принять", "Хорошо", "Понятно", "Accept all"))


async def wait_if_blocked(page, log: Callable[[str], None]) -> None:
    try:
        content = str(await page.evaluate("() => (document.body && document.body.innerText) || ''")).lower()
    except Exception:
        return
    if any(marker in content for marker in BLOCK_MARKERS):
        log("Авито показал проверку (капча / доступ ограничен).")
        log("Решите проверку в открытом окне Chrome, программа подождет 90 секунд.")
        await page.wait_for_timeout(90_000)


async def extract_listing_items(page) -> list[dict]:
    return await page.evaluate(
        """
        () => {
          const pickText = (root, selectors) => {
            for (const selector of selectors) {
              const el = root.querySelector(selector);
              const text = el && (el.innerText || el.textContent || '').trim();
              if (text) return text;
            }
            return '';
          };
          const cards = Array.from(document.querySelectorAll('[data-marker="item"]'));
          return cards.map(card => {
            const link = card.querySelector('a[data-marker="item-title"], a[itemprop="url"], a[href*="_"]');
            const priceMeta = card.querySelector('meta[itemprop="price"]');
            return {
              title: pickText(card, ['[itemprop="name"]', 'a[data-marker="item-title"]', 'h3']),
              href: link ? (link.href || '') : '',
              price: priceMeta ? priceMeta.getAttribute('content') || '' :
                pickText(card, ['[data-marker="item-price"]', '[itemprop="offers"]', 'p']),
              location: pickText(card, [
                '[data-marker="item-address"]',
                '[class*="geo-root"]',
                '[class*="geo-address"]'
              ]),
              description: pickText(card, [
                '[data-marker="item-descriptionText"]',
                '[class*="item-description"]',
                'meta[itemprop="description"]'
              ]),
              published: pickText(card, ['[data-marker="item-date"]', '[class*="date-text"]'])
            };
          }).filter(item => item.href);
        }
        """
    )


async def enrich_from_item_page(page, item: AvitoItem, log: Callable[[str], None]) -> None:
    try:
        await page.goto(item.url, wait_until="domcontentloaded", timeout=60_000)
        await page.wait_for_timeout(2500)
        await wait_if_blocked(page, log)
        data = await page.evaluate(
            """
            () => {
              const pick = selectors => {
                for (const selector of selectors) {
                  const el = document.querySelector(selector);
                  const text = el && (el.innerText || el.textContent || '').trim();
                  if (text) return text;
                }
                return '';
              };
              return {
                seller: pick([
                  '[data-marker="seller-info/name"]',
                  '[data-marker="seller-link/linkText"]',
                  '[data-marker="seller-info/label"]'
                ]),
                description: pick([
                  '[data-marker="item-view/item-description"]',
                  '[itemprop="description"]'
                ]),
                address: pick([
                  '[data-marker="item-view/item-address"]',
                  '[itemprop="address"]'
                ])
              };
            }
            """
        )
        if clean_text(data.get("seller")):
            item.seller_name = clean_text(data.get("seller"))
        if clean_text(data.get("description")):
            item.description = clean_text(data.get("description"))[:900]
        if clean_text(data.get("address")):
            item.location = clean_text(data.get("address"))
    except PlaywrightTimeoutError:
        log("  ! таймаут карточки, оставляю данные из выдачи")
    except Exception as exc:
        log(f"  ! ошибка карточки: {exc}")


async def scrape_avito(
    *,
    query: str,
    city: str,
    max_results: int,
    max_pages: int,
    deep: bool,
    visible: bool,
    output_dir: Path,
    log: Callable[[str], None],
    item_callback: Optional[Callable[[AvitoItem], None]] = None,
) -> tuple[list[AvitoItem], dict[str, Path]]:
    chrome = find_chrome()
    DEFAULT_PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    log(f"Chrome: {chrome}")
    log(f"Открываю Авито: {query} / {city}")
    items: list[AvitoItem] = []
    seen: set[str] = set()

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            str(DEFAULT_PROFILE_DIR),
            executable_path=str(chrome),
            headless=not visible,
            viewport={"width": 1360, "height": 900},
            locale="ru-RU",
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-first-run",
                "--no-default-browser-check",
            ],
        )
        page = context.pages[0] if context.pages else await context.new_page()
        try:
            stagnant_pages = 0
            for page_no in range(1, max_pages + 1):
                if len(items) >= max_results:
                    break
                url = build_search_url(query, city, page_no)
                log(f"Страница выдачи {page_no}: {url}")
                try:
                    await page.goto(url, wait_until="domcontentloaded", timeout=70_000)
                except PlaywrightTimeoutError:
                    log("  ! таймаут страницы выдачи, пропускаю")
                    continue
                await page.wait_for_timeout(4500)
                if page_no == 1:
                    await accept_cookies(page)
                await wait_if_blocked(page, log)

                raw_items = await extract_listing_items(page)
                new_count = 0
                for raw in raw_items:
                    clean_url = canonical_item_url(raw.get("href") or "")
                    if not clean_url or item_key(clean_url) in seen:
                        continue
                    seen.add(item_key(clean_url))
                    item = AvitoItem(
                        title=clean_text(raw.get("title")),
                        price=normalize_price(str(raw.get("price") or "")),
                        url=clean_url,
                        location=clean_text(raw.get("location")),
                        description=clean_text(raw.get("description"))[:900],
                        published=clean_text(raw.get("published")),
                        source_query=query,
                        source_city=city,
                        collected_at=time.strftime("%Y-%m-%d %H:%M:%S"),
                    )
                    if not item.title:
                        continue
                    items.append(item)
                    new_count += 1
                    if not deep:
                        if item_callback:
                            item_callback(item)
                        log(f"  + {item.title} | {item.price or 'без цены'}")
                    if len(items) >= max_results:
                        break

                log(f"Собрано объявлений: {len(items)}")
                if new_count == 0:
                    stagnant_pages += 1
                    if stagnant_pages >= 2:
                        log("Новых объявлений нет, останавливаю листание.")
                        break
                else:
                    stagnant_pages = 0

            if deep:
                log(f"Открываю карточки объявлений: {len(items)}")
                for index, item in enumerate(items, start=1):
                    log(f"[{index}/{len(items)}] {item.url}")
                    await enrich_from_item_page(page, item, log)
                    if item_callback:
                        item_callback(item)
                    await page.wait_for_timeout(1200)
        finally:
            await context.close()

    paths = write_outputs(items, output_dir)
    log(f"Готово. Объявлений: {len(items)}")
    log(f"Excel: {paths['xlsx']}")
    return items, paths


def write_outputs(items: list[AvitoItem], output_dir: Path) -> dict[str, Path]:
    stamp = time.strftime("%Y%m%d_%H%M%S")
    csv_path = output_dir / f"avito_items_{stamp}.csv"
    xlsx_path = output_dir / f"avito_items_{stamp}.xlsx"
    json_path = output_dir / f"avito_items_{stamp}.json"
    fields = list(AvitoItem.__dataclass_fields__.keys())

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for item in items:
            writer.writerow(asdict(item))

    json_path.write_text(json.dumps([asdict(item) for item in items], ensure_ascii=False, indent=2), encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "Avito"
    ws.append(fields)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")
    for item in items:
        ws.append([getattr(item, field) for field in fields])
    widths = {
        "A": 46,
        "B": 14,
        "C": 56,
        "D": 30,
        "E": 26,
        "F": 60,
        "G": 18,
        "H": 22,
        "I": 16,
        "J": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(xlsx_path)
    return {"csv": csv_path, "xlsx": xlsx_path, "json": json_path}


class AvitoParserApp:
    def __init__(self) -> None:
        self.root = tk.Tk()
        self.root.title(APP_TITLE)
        self.root.geometry("1180x760")
        self.queue: queue.Queue[tuple[str, object]] = queue.Queue()
        self.worker: Optional[threading.Thread] = None
        self.last_output_dir = DEFAULT_OUTPUT_DIR
        self._build_ui()
        self.root.after(150, self._drain_queue)

    def _build_ui(self) -> None:
        self.root.columnconfigure(1, weight=1)
        self.root.rowconfigure(1, weight=1)

        left = ttk.Frame(self.root, padding=12)
        left.grid(row=0, column=0, rowspan=2, sticky="nsew")
        left.columnconfigure(0, weight=1)

        title = ttk.Label(left, text="Avito Parser", font=("Segoe UI", 16, "bold"))
        title.grid(row=0, column=0, sticky="w", pady=(0, 12))

        self.query_var = tk.StringVar(value="пастила")
        self.city_var = tk.StringVar(value="Москва")
        self.limit_var = tk.IntVar(value=100)
        self.pages_var = tk.IntVar(value=10)
        self.deep_var = tk.BooleanVar(value=False)
        self.visible_var = tk.BooleanVar(value=True)
        self.output_var = tk.StringVar(value=str(DEFAULT_OUTPUT_DIR))

        ttk.Label(left, text="Что искать").grid(row=1, column=0, sticky="w")
        ttk.Entry(left, textvariable=self.query_var, width=34).grid(row=2, column=0, sticky="ew", pady=(2, 10))

        ttk.Label(left, text="Город").grid(row=3, column=0, sticky="w")
        city = ttk.Combobox(left, textvariable=self.city_var, values=[name.title() for name in CITY_SLUGS], width=32)
        city.grid(row=4, column=0, sticky="ew", pady=(2, 10))

        ttk.Label(left, text="Лимит объявлений").grid(row=5, column=0, sticky="w")
        ttk.Spinbox(left, from_=1, to=2000, textvariable=self.limit_var, width=10).grid(row=6, column=0, sticky="w", pady=(2, 10))

        ttk.Label(left, text="Максимум страниц выдачи").grid(row=7, column=0, sticky="w")
        ttk.Spinbox(left, from_=1, to=100, textvariable=self.pages_var, width=10).grid(row=8, column=0, sticky="w", pady=(2, 10))

        ttk.Checkbutton(
            left,
            text="Открывать каждую карточку\n(медленнее: продавец и описание)",
            variable=self.deep_var,
        ).grid(row=9, column=0, sticky="w", pady=(0, 6))

        ttk.Checkbutton(left, text="Показывать Chrome во время сбора", variable=self.visible_var).grid(
            row=10, column=0, sticky="w", pady=(0, 10)
        )

        ttk.Label(left, text="Папка результатов").grid(row=11, column=0, sticky="w")
        ttk.Entry(left, textvariable=self.output_var, width=34).grid(row=12, column=0, sticky="ew", pady=(2, 6))
        ttk.Button(left, text="Выбрать папку", command=self._choose_output_dir).grid(row=13, column=0, sticky="ew", pady=(0, 10))

        self.start_button = ttk.Button(left, text="Запустить сбор", command=self._start)
        self.start_button.grid(row=14, column=0, sticky="ew", pady=(4, 8))
        ttk.Button(left, text="Открыть папку результатов", command=self._open_output_dir).grid(row=15, column=0, sticky="ew")

        hint = ttk.Label(
            left,
            text="Программа открывает обычный Google Chrome,\nлистает выдачу Авито и сохраняет Excel.\nЕсли Авито покажет капчу — решите ее\nв окне Chrome, сбор продолжится сам.",
            foreground="#335b9f",
            justify="left",
        )
        hint.grid(row=16, column=0, sticky="w", pady=(18, 0))

        right = ttk.Frame(self.root, padding=(0, 12, 12, 12))
        right.grid(row=0, column=1, rowspan=2, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=3)
        right.rowconfigure(1, weight=2)

        columns = ("title", "price", "location", "published", "url")
        self.table = ttk.Treeview(right, columns=columns, show="headings", height=16)
        for column, text, width in (
            ("title", "Объявление", 280),
            ("price", "Цена", 100),
            ("location", "Локация", 180),
            ("published", "Дата", 110),
            ("url", "Ссылка", 320),
        ):
            self.table.heading(column, text=text)
            self.table.column(column, width=width, anchor="w")
        self.table.grid(row=0, column=0, sticky="nsew")
        table_scroll = ttk.Scrollbar(right, orient="vertical", command=self.table.yview)
        table_scroll.grid(row=0, column=1, sticky="ns")
        self.table.configure(yscrollcommand=table_scroll.set)

        log_frame = ttk.LabelFrame(right, text="Лог")
        log_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(10, 0))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        self.log_text = tk.Text(log_frame, height=10, bg="#101827", fg="#e6eefb", insertbackground="#e6eefb")
        self.log_text.grid(row=0, column=0, sticky="nsew")
        log_scroll = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        log_scroll.grid(row=0, column=1, sticky="ns")
        self.log_text.configure(yscrollcommand=log_scroll.set)

        self.status_var = tk.StringVar(value="Готов к работе")
        status = ttk.Label(self.root, textvariable=self.status_var, padding=(12, 6), foreground="#1b4f9c")
        status.grid(row=2, column=0, columnspan=2, sticky="ew")

    def _choose_output_dir(self) -> None:
        selected = filedialog.askdirectory(initialdir=self.output_var.get() or str(DEFAULT_OUTPUT_DIR))
        if selected:
            self.output_var.set(selected)

    def _open_output_dir(self) -> None:
        path = Path(self.output_var.get() or self.last_output_dir)
        path.mkdir(parents=True, exist_ok=True)
        subprocess.Popen(["explorer", str(path)])

    def _start(self) -> None:
        if self.worker and self.worker.is_alive():
            messagebox.showinfo(APP_TITLE, "Сбор уже идет.")
            return
        query = clean_text(self.query_var.get())
        city = clean_text(self.city_var.get())
        if not query:
            messagebox.showwarning(APP_TITLE, "Введите запрос: например, пастила или велосипед.")
            return
        max_results = max(1, int(self.limit_var.get() or 1))
        max_pages = max(1, int(self.pages_var.get() or 1))
        output_dir = Path(self.output_var.get() or DEFAULT_OUTPUT_DIR)
        self.last_output_dir = output_dir
        for row in self.table.get_children():
            self.table.delete(row)
        self.log_text.delete("1.0", "end")
        self.start_button.configure(state="disabled")
        self.status_var.set("Сбор запущен")

        def worker() -> None:
            try:
                asyncio.run(
                    scrape_avito(
                        query=query,
                        city=city,
                        max_results=max_results,
                        max_pages=max_pages,
                        deep=bool(self.deep_var.get()),
                        visible=bool(self.visible_var.get()),
                        output_dir=output_dir,
                        log=lambda message: self.queue.put(("log", message)),
                        item_callback=lambda item: self.queue.put(("item", item)),
                    )
                )
                self.queue.put(("done", output_dir))
            except Exception as exc:
                self.queue.put(("error", str(exc)))

        self.worker = threading.Thread(target=worker, daemon=True)
        self.worker.start()

    def _drain_queue(self) -> None:
        try:
            while True:
                event, payload = self.queue.get_nowait()
                if event == "log":
                    self._log(str(payload))
                elif event == "item" and isinstance(payload, AvitoItem):
                    self.table.insert(
                        "",
                        "end",
                        values=(payload.title, payload.price, payload.location, payload.published, payload.url),
                    )
                elif event == "done":
                    self.start_button.configure(state="normal")
                    self.status_var.set("Готово")
                    self._log("Сбор завершен.")
                    messagebox.showinfo(APP_TITLE, "Сбор завершен. Excel и CSV сохранены в папку результатов.")
                elif event == "error":
                    self.start_button.configure(state="normal")
                    self.status_var.set("Ошибка")
                    self._log(f"Ошибка: {payload}")
                    messagebox.showerror(APP_TITLE, str(payload))
        except queue.Empty:
            pass
        self.root.after(150, self._drain_queue)

    def _log(self, message: str) -> None:
        self.log_text.insert("end", f"{time.strftime('%H:%M:%S')} | {message}\n")
        self.log_text.see("end")

    def run(self) -> None:
        self.root.mainloop()


def parse_args(argv: Optional[list[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Desktop Avito parser")
    parser.add_argument("--cli", action="store_true", help="Run without desktop UI")
    parser.add_argument("--query", default="пастила")
    parser.add_argument("--city", default="Москва")
    parser.add_argument("--limit", type=int, default=100)
    parser.add_argument("--pages", type=int, default=10)
    parser.add_argument("--deep", action="store_true", help="Open every item page for seller info")
    parser.add_argument("--headless", action="store_true")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args(argv)


def main(argv: Optional[list[str]] = None) -> int:
    args = parse_args(argv)
    if args.cli:
        for stream in (sys.stdout, sys.stderr):
            try:
                stream.reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass
        items, paths = asyncio.run(
            scrape_avito(
                query=args.query,
                city=args.city,
                max_results=args.limit,
                max_pages=args.pages,
                deep=args.deep,
                visible=not args.headless,
                output_dir=args.output_dir,
                log=print,
            )
        )
        print(f"Saved {len(items)} items")
        print(paths["xlsx"])
        return 0
    app = AvitoParserApp()
    app.run()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

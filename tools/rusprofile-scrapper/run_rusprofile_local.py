#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Local Rusprofile runner that exports to files instead of MySQL."""

from __future__ import annotations

import argparse
import csv
import json
import sys
import time
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from scrapy.crawler import CrawlerProcess


BASE_DIR = Path(__file__).resolve().parent
from rusprofile_modern_spider import ModernRusprofileSpider  # noqa: E402


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run rusprofile scraper locally and export CSV/JSON/XLSX.")
    parser.add_argument("ids", nargs="+", help="Rusprofile code IDs, e.g. 89220")
    parser.add_argument("--output-dir", type=Path, default=BASE_DIR / "results")
    parser.add_argument("--concurrency", type=int, default=2)
    parser.add_argument("--delay", type=float, default=3.0)
    parser.add_argument("--max-items", type=int, default=100, help="Stop after this many companies")
    return parser.parse_args()


def write_xlsx(rows: list[dict[str, Any]], path: Path) -> None:
    fields = [
        "name",
        "inn",
        "ogrn",
        "okpo",
        "current_status",
        "reg_date",
        "auth_capital",
        "address",
        "source_url",
        "source_code_url",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Rusprofile"
    ws.append(fields)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")
    for row in rows:
        ws.append([row.get(field, "") for field in fields])
    widths = {"A": 44, "B": 16, "C": 18, "D": 14, "E": 24, "F": 14, "G": 16, "H": 60, "I": 48, "J": 48}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main() -> int:
    args = parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    csv_path = args.output_dir / f"rusprofile_{stamp}.csv"
    json_path = args.output_dir / f"rusprofile_{stamp}.json"
    xlsx_path = args.output_dir / f"rusprofile_{stamp}.xlsx"

    settings = {
        "LOG_LEVEL": "INFO",
        "COOKIES_ENABLED": False,
        "CONCURRENT_REQUESTS": max(1, args.concurrency),
        "DOWNLOAD_DELAY": max(0.5, args.delay),
        "DOWNLOAD_TIMEOUT": 20,
        "RETRY_TIMES": 4,
        "AUTOTHROTTLE_ENABLED": True,
        "AUTOTHROTTLE_START_DELAY": max(0.5, args.delay),
        "AUTOTHROTTLE_MAX_DELAY": 30,
        "AUTOTHROTTLE_TARGET_CONCURRENCY": 1.0,
        "CLOSESPIDER_ITEMCOUNT": max(1, args.max_items),
        "FEED_EXPORT_ENCODING": "utf-8",
        "USER_AGENT": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
        ),
        "ITEM_PIPELINES": {},
        "DOWNLOADER_MIDDLEWARES": {},
        "FEEDS": {
            str(csv_path): {"format": "csv", "encoding": "utf-8-sig", "overwrite": True},
            str(json_path): {"format": "json", "encoding": "utf-8", "indent": 2, "overwrite": True},
        },
    }

    process = CrawlerProcess(settings=settings)
    process.crawl(ModernRusprofileSpider, ids=args.ids)
    process.start()

    rows: list[dict[str, Any]] = []
    if csv_path.exists():
        with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.DictReader(f))
    write_xlsx(rows, xlsx_path)

    print(f"Saved rows: {len(rows)}")
    print(f"CSV:  {csv_path}")
    print(f"JSON: {json_path}")
    print(f"XLSX: {xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
